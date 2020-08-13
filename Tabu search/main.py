import tsplib95
from pprint import pprint
import random as rnd
import math
import copy
from random import randrange
import matplotlib.pyplot as plt
import time
import openpyxl
import os
import statistics
import numpy as np

NUM_ITERATIONS = 500
DEBUG = False
graph = None
EXEl_WRITE = False
dependencies = []
INIT_HEURISTIC = True


class Edge(object):

    def __init__(self, vertices, weight):
        self.vertices = vertices
        self.weight = weight

    def __str__(self):
        return str(self.vertices) + "->" + str(self.weight)
        # return str(self.weight)
        # return str(self.vertices)

    def __repr__(self):
        return str(self)


class Graph(object):

    def __init__(self, problem):
        self.edges = []
        self.dependencies = []
        self.dimension = problem.dimension
        problemEgdes = list(problem.get_edges())
        problemWeights = problem.edge_weights[1:]

        for i in range(len(problemEgdes)):
            self.edges.append(Edge(problemEgdes[i], problemWeights[i]))


def calculateDependencies(problem):
    dependencies = []
    edgeWeights = problem.edge_weights[1:]

    for i in range(problem.dimension):
        dependencies.append(list())
        for j in range(graph.dimension):
            if(edgeWeights[(i*problem.dimension)+j] == -1):
                dependencies[i].append(j)
    return dependencies


def fpp3exchange(problem, deps, solution):
    dimension = problem.dimension
    edgeWeights = problem.edge_weights[1:]

    solutions = []
    for it in range(int(dimension/2)):

        # h < dimension-2
        h = randrange(0, dimension-3)
        i = h + 1
        leftPath = []
        leftPathLen = randrange(1, int(dimension-i))
        leftPath.extend(solution[i:i+leftPathLen])

        i += leftPathLen
        end = False
        rightPath = []
        for j in range(i, len(solution)):

            for dep in deps[solution[j]]:
                if dep != 0 and dep in leftPath:
                    end = True
                    break

            # terminate the progress
            if end:
                break
            # add j to right path
            else:
                rightPath.append(solution[j])

        if(len(rightPath) != 0):
            sol = solution[0:h+1]
            sol.extend(rightPath)
            sol.extend(leftPath)
            sol.extend(solution[len(sol):])

            # creating action for saving in tabu list
            action = (solution[h], solution[i], solution[j])

            solutions.append((sol, costFunction(problem, sol), action))

    solutions.sort(key=lambda x: x[1])
    if len(solutions) != 0:
        return solutions[0]
    else:
        return None


def bpp3exchange(problem, deps, solution):
    dimension = problem.dimension
    edgeWeights = problem.edge_weights[1:]

    solutions = []
    for it in range(int(dimension/2)):
        h = randrange(3, dimension)
        i = h - 1
        rightPath = []
        rightPathLen = randrange(1, i+1)
        rightPath.extend(solution[i-rightPathLen+1:i+1])
        rightDeps = []

        for node in rightPath:
            rightDeps.extend(deps[node])

        i -= rightPathLen

        leftPath = []
        for j in range(i, 0, -1):

            # add j to left path
            if solution[j] not in rightDeps:
                leftPath.insert(0, solution[j])
            else:
                break

        if(len(leftPath) != 0):
            sol = solution[h:]
            sol = leftPath + sol
            sol = rightPath + sol
            sol = solution[:dimension - len(sol)] + sol

            # creating action for saving in tabu list
            action = (solution[h], solution[i], solution[j])

            solutions.append((sol, costFunction(problem, sol), action))

    solutions.sort(key=lambda x: x[1])
    if len(solutions) != 0:
        return solutions[0]
    else:
        return None


def costFunction(problem, solution):

    weight = 0
    edgeWeights = problem.edge_weights[1:]
    sol = copy.deepcopy(solution)

    while(len(sol) > 1):
        src = sol.pop(0)
        dest = sol[0]
        w = edgeWeights[(src*problem.dimension)+dest]
        weight += w

    return weight


def initialStart(graph, deps):
    solution = []

    # array of all nodes dependencies
    dependencies = copy.deepcopy(deps)

    while(len(solution) < graph.dimension):
        for i in range(graph.dimension):
            if(INIT_HEURISTIC):
                src = 0
                if len(solution) != 0:
                    src = solution[-1]

                candidates = []

                # determining feasible destinations
                result = [i for i in range(
                    len(dependencies)) if len(dependencies[i]) == 0]

                # store feasible destinations and edge weight from source to them
                candidates = [
                    (i, graph.edges[(src*graph.dimension) + i].weight)
                    for i in result if i not in solution]

                # gready manner : sort from best to worst
                candidates = sorted(candidates, key=lambda tup: tup[1])

                solution.append(candidates[0][0])

                # updates node dependencies
                for dep in dependencies:
                    if(candidates[0][0] in dep):
                        dep.remove(candidates[0][0])

            else:
                if(len(dependencies[i]) == 0 and not(i in solution)):
                    solution.append(i)

                    # updates node dependencies
                    for dep in dependencies:
                        if(i in dep):
                            dep.remove(i)

    return solution


def improveSolution(problem, dependencies, solution, tabuList,
                    MAX_MEM_DEPTH):
    (state, cost) = solution

    new_solutions = []
    new_solutions1 = fpp3exchange(problem, dependencies, state)
    new_solutions2 = bpp3exchange(problem, dependencies, state)

    if new_solutions1 != None:
        new_solutions.append(new_solutions1)
    if new_solutions2 != None:
        new_solutions.append(new_solutions2)

    if len(new_solutions) != 0:
        new_solutions.sort(key=lambda x: x[1])

        tabuActions = [x[0] for x in tabuList]

        if len(tabuList) != 0:
            minFrequentActions = min(tabuList, key=lambda ac: ac[1])
        else:
            minFrequentActions = []

        # checking feasiblity based on tabu list
        for sol in new_solutions:
            action = sol[2]

            if action not in tabuActions:
                # add action to tabu list
                tabuList.append((action, MAX_MEM_DEPTH))

                # returining solution and its cost
                return (sol[0], sol[1])


        # if all actions was tabu return previous solution
        return (state, cost)
    else:
        return (state, cost)


def updateTabuList(tabuList, TABU_LIST_SIZE):

    if len(tabuList) > TABU_LIST_SIZE:
        del tabuList[0]

    # updating all actions duration
    tabuList = [(x[0], x[1]-1) for x in tabuList]

    # aspiration condition
    tabuList = [x for x in tabuList if x[1] != 0]


    return tabuList


def TabuSearch(problem, initialStart, costFunction, improveSolution, updateTabuList, maxsteps=1000,
               TABU_LIST_SIZE=10, MAX_MEM_DEPTH=10, MAX_ACTION_NUM=100, debug=True):

    tabuList = []
    history = []

    # initialing best solution
    state = initialStart(graph, dependencies)
    cost = costFunction(problem, state)
    bestSolution = (state, cost)

    for step in range(maxsteps):

        # improving the best solution
        (bestSolution) = improveSolution(problem, dependencies,
                                         bestSolution, tabuList, MAX_MEM_DEPTH)

        # updating tabu list
        tabuList = updateTabuList(tabuList, TABU_LIST_SIZE)

        # updating search history
        history.append(bestSolution[1])

        # if DEBUG:
        #     print('\nstep:', step, '\t bestSolution:', bestSolution)

    return bestSolution, history


def plotResult(costs):

    plt.plot(list(range(len(costs))), costs, '-', color="gray",
             label='algorithm progress', linewidth=1.5)
    plt.xlabel('best solution')
    plt.ylabel('iteration')
    plt.show()


def plotResultPerX(x, costs):

    print("tabuSizes:", x, "costs:", costs)
    plt.plot(x, costs, '-', color="gray",
             label='algorithm progress', linewidth=1.5)
    plt.xlabel('Action expiration time as coefficient of n')
    plt.ylabel('average cost')
    plt.show()


def printResult(answers):

    minAns = min(answers, key=lambda t: t[1])
    maxAns = max(answers, key=lambda t: t[1])
    variance = math.sqrt(np.var([ans[1] for ans in answers]))

    print("\nbest[0:10]=", minAns[0][0:10], "\tmin cost:", minAns[1])
    print("worst[0:10]=", maxAns[0][0:10], "\tmax cost:",
          max(answers, key=lambda t: t[1])[1])
    print("\naverage cost:", sum(ans[1] for ans in answers)/len(answers))
    print("\nvariance of costs:", variance)

    print("\nmin time:", min(answers, key=lambda t: t[2])[2])
    print("avg time:", str(sum(float(ans[2])
                               for ans in answers)/len(answers))[0:6])
    print("max time:", max(answers, key=lambda t: t[2])[2])


def writeResultToExel(file_name, answers, myRow):
    minCost = min(answers, key=lambda t: t[1])[1]
    maxCost = max(answers, key=lambda t: t[1])[1]
    avgCost = sum(ans[1] for ans in answers)/len(answers)
    costVariance = math.sqrt(np.var([ans[1]for ans in answers]))

    minTime = min(answers, key=lambda t: t[2])[2]
    maxTime = max(answers, key=lambda t: t[2])[2]
    avgTime = str(sum(float(ans[2])for ans in answers)/len(answers))[0:6]

    wbkName = 'Results.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        myCol = 4

        wks.cell(row=myRow, column=1).value = file_name

        wks.cell(row=myRow, column=myCol).value = minCost
        wks.cell(row=myRow, column=myCol+1).value = avgCost
        wks.cell(row=myRow, column=myCol+2).value = maxCost
        wks.cell(row=myRow, column=myCol+3).value = costVariance

        wks.cell(row=myRow, column=myCol+4).value = minTime
        wks.cell(row=myRow, column=myCol+5).value = avgTime
        wks.cell(row=myRow, column=myCol+6).value = maxTime

    wbk.save(wbkName)
    wbk.close

def plotBoxDiagram(answers):
    costs = [ans[1] for ans in answers]
    plt.boxplot(costs)
    plt.show()

if __name__ == '__main__':

    myRow = 2
    # for root, directories, filenames in os.walk("instances/"):
    #     for filename in filenames:
    #         file = os.path.join(root, filename)
    #         problem = tsplib95.load_problem(str(file))
    problem = tsplib95.load_problem("instances/E/prob.7.65.sop")

    graph = Graph(problem)
    dependencies = calculateDependencies(problem)
    answers = []

    # initialing tabu list size
    TABU_LIST_SIZE = int(problem.dimension/14)
    MAX_MEM_DEPTH = int(problem.dimension/16)

    n = problem.dimension

    print("\ninstance:", problem.name, "TABU_LIST_SIZE:",
        TABU_LIST_SIZE, "MAX_MEM_DEPTH:", MAX_MEM_DEPTH, "\n")

    for i in range(10):
        start = time.time()

        (state, cost), history = TabuSearch(problem, initialStart, costFunction,
                                            improveSolution, updateTabuList,
                                            NUM_ITERATIONS, TABU_LIST_SIZE,
                                            MAX_MEM_DEPTH, DEBUG)

        
        duration = str(time.time() - start)[0:6]
        answers.append((state, cost, duration))
        print('time:',duration , "cost:",cost)

    printResult(answers)
    plotBoxDiagram(answers)
    if EXEl_WRITE:
        writeResultToExel(filename, answers, myRow)
        myRow += 1
