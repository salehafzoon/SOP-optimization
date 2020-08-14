import matplotlib.pyplot as plt
from random import randrange
from pprint import pprint
import random as rn
import numpy as np
import tsplib95
import openpyxl
import math
import copy
import time
import os

LINEAR = 'linear'
LOG = 'logarithmic'
EXP = 'exponential'

INIT_HEURISTIC = True
NUM_ITERATIONS = 500
dependencies = []
EPSILON = 1e-323
TEMP_MODE = EXP
DEBUG = False
graph = None
START_T = 1
T = START_T
ALPHA = 0.9

EXCEl_WRITE = False

class Edge(object):

    def __init__(self, vertices, weight):
        self.vertices = vertices
        self.weight = weight

    def __str__(self):
        return str(self.vertices) + "->" + str(self.weight)
    def __repr__(self):
        return str(self)


class Graph(object):

    def __init__(self, problem):
        self.edges = []
        self.dependencies = []
        self.dimension = problem.dimension
        problemEgdes = list(problem.get_edges())
        problemWeights = problem.edge_weights[1:]

        for i in range(len(problemEgdes)):
            self.edges.append(Edge(problemEgdes[i], problemWeights[i]))


def calculateDependencies(problem):
    dependencies = []
    edgeWeights = problem.edge_weights[1:]

    for i in range(problem.dimension):
        dependencies.append(list())
        for j in range(graph.dimension):
            if(edgeWeights[(i*problem.dimension)+j] == -1):
                dependencies[i].append(j)
    return dependencies


def fpp3exchange(problem, deps, solution):
    dimension = problem.dimension
    edgeWeights = problem.edge_weights[1:]

    solutions = []
    for it in range(int(dimension/2)):
        h = randrange(0, dimension-3)
        i = h + 1
        leftPath = []
        leftPathLen = randrange(1, int(dimension-i))
        leftPath.extend(solution[i:i+leftPathLen])

        i += leftPathLen
        end = False
        rightPath = []
        for j in range(i, len(solution)):

            for dep in deps[solution[j]]:
                if dep != 0 and dep in leftPath:
                    end = True
                    break

            # terminate the progress
            if end:
                break
            # add j to right path
            else:
                rightPath.append(solution[j])

        if(len(rightPath) != 0):
            sol = solution[0:h+1]
            sol.extend(rightPath)
            sol.extend(leftPath)
            sol.extend(solution[len(sol):])
            solutions.append((sol, cost_function(problem, sol)))

    solutions.sort(key=lambda x: x[1])
    if len(solutions) != 0:
        return solutions[0]
    else:
        return None


def bpp3exchange(problem, deps, solution):
    dimension = problem.dimension
    edgeWeights = problem.edge_weights[1:]

    solutions = []
    for it in range(int(dimension/2)):
        h = randrange(3, dimension)
        i = h - 1
        rightPath = []
        rightPathLen = randrange(1, i+1)
        rightPath.extend(solution[i-rightPathLen+1:i+1])
        rightDeps = []

        for node in rightPath:
            rightDeps.extend(deps[node])

        i -= rightPathLen

        leftPath = []
        for j in range(i, 0, -1):

            # add j to left path
            if solution[j] not in rightDeps:
                leftPath.insert(0, solution[j])
            else:
                break

        if(len(leftPath) != 0):
            sol = solution[h:]
            sol = leftPath + sol
            sol = rightPath + sol
            sol = solution[:dimension - len(sol)] + sol
            solutions.append((sol, cost_function(problem, sol)))

    solutions.sort(key=lambda x: x[1])
    if len(solutions) != 0:
        return solutions[0]
    else:
        return None


def random_start(graph, deps):
    solution = []
    dependencies = copy.deepcopy(deps)

    while(len(solution) < graph.dimension):
        for i in range(graph.dimension):
            if(INIT_HEURISTIC):
                src = 0
                if len(solution) != 0:
                    src = solution[-1]

                if len(solution) == 7:
                    pass

                candidates = []

                result = [i for i in range(
                    len(dependencies)) if len(dependencies[i]) == 0]

                candidates = [
                    (i, graph.edges[(src*graph.dimension) + i].weight)
                    for i in result if i not in solution]

                candidates = sorted(candidates, key=lambda tup: tup[1])

                solution.append(candidates[0][0])

                for dep in dependencies:
                    if(candidates[0][0] in dep):
                        dep.remove(candidates[0][0])

            else:
                if(len(dependencies[i]) == 0 and not(i in solution)):
                    solution.append(i)
                    for dep in dependencies:
                        if(i in dep):
                            dep.remove(i)

    return solution


def cost_function(problem, solution):

    weight = 0
    edgeWeights = problem.edge_weights[1:]
    sol = copy.deepcopy(solution)

    while(len(sol) > 1):
        src = sol.pop(0)
        dest = sol[0]
        w = edgeWeights[(src*problem.dimension)+dest]
        weight += w

    return weight


def acceptance_probability(cost, new_cost, temperature):
    if new_cost < cost:
        return 1
    else:
        p = math.exp(- (new_cost - cost) / temperature)
        return p


def get_neighbour(problem, dependencies, state, cost):
    new_states = []
    new_state1 = fpp3exchange(problem, dependencies, state)
    new_state2 = bpp3exchange(problem, dependencies, state)

    if new_state1 != None:
        new_states.append(new_state1)
    if new_state2 != None:
        new_states.append(new_state2)

    if len(new_states) != 0:
        new_states.sort(key=lambda x: x[1])
        return new_states[0]

    else:
        return (state, cost)


def updateTemperature(step):
    global T
    if TEMP_MODE == LINEAR:
        return ALPHA * T
    elif TEMP_MODE == LOG:
        return START_T / math.log(step+2)
    elif TEMP_MODE == EXP:
        return math.exp(-ALPHA * step+1)*START_T


def annealing(problem,random_start, cost_function, random_neighbour,
              acceptance, updateTemperature, maxsteps=1000, debug=True):

    global T
    state = random_start(graph, dependencies)
    cost = cost_function(problem, state)
    states, costs = [state], [cost]
    for step in range(maxsteps):
        (new_state, new_cost) = get_neighbour(
            problem, dependencies, state, cost)
        if debug:

            print('step:', step, '\t T:', T, '\t new_cost:', new_cost)

        if acceptance_probability(cost, new_cost, T) > rn.random():
            state, cost = new_state, new_cost
            states.append(state)
            costs.append(cost)

        T = updateTemperature(step)
        if T == 0.0:
            T = EPSILON

    return new_state, new_cost, states, costs


def printResult(answers):

    minAns = min(answers, key=lambda t: t[1])
    maxAns = max(answers, key=lambda t: t[1])
    variance = round(math.sqrt(np.var([ans[1]for ans in answers])) , 3)

    print("-------------------")
    print("\nbest[0:10]=", minAns[0][0:10], "\tmin cost:", minAns[1])
    print("worst[0:10]=", maxAns[0][0:10], "\tmax cost:",
          max(answers, key=lambda t: t[1])[1])
    print("\naverage cost:", sum(ans[1] for ans in answers)/len(answers))
    print("variance of costs:", variance)

    print("\nmin time:", min(answers, key=lambda t: t[2])[2])
    print("avg time:", str(sum(float(ans[2])
                               for ans in answers)/len(answers))[0:6])
    print("max time:", max(answers, key=lambda t: t[2])[2])
    print("-------------------")
    

def writeResultToExcel(file_name, answers, myRow):
    minCost = min(answers, key=lambda t: t[1])[1]
    maxCost = max(answers, key=lambda t: t[1])[1]
    avgCost = sum(ans[1] for ans in answers)/len(answers)
    costVariance = round(math.sqrt(np.var([ans[1]for ans in answers])) , 3)

    minTime = min(answers, key=lambda t: t[2])[2]
    maxTime = max(answers, key=lambda t: t[2])[2]
    avgTime = str(sum(float(ans[2])for ans in answers)/len(answers))[0:6]

    wbkName = './SA/Results.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    for wks in wbk.worksheets:
        myCol = 4

        wks.cell(row=myRow, column=1).value = file_name

        wks.cell(row=myRow, column=myCol).value = minCost
        wks.cell(row=myRow, column=myCol+1).value = avgCost
        wks.cell(row=myRow, column=myCol+2).value = maxCost
        wks.cell(row=myRow, column=myCol+3).value = costVariance

        wks.cell(row=myRow, column=myCol+4).value = minTime
        wks.cell(row=myRow, column=myCol+5).value = avgTime
        wks.cell(row=myRow, column=myCol+6).value = maxTime

    wbk.save(wbkName)
    wbk.close


if __name__ == '__main__':

    myRow = 2
    for root, directories, filenames in os.walk("./instances/H"):
        for filename in filenames:
            file = os.path.join(root, filename)
            problem = tsplib95.load_problem(str(file))

            graph = Graph(problem)
            dependencies = calculateDependencies(problem)
            answers = []

            print("\ninstance:", problem.name, "\tTEMP_MODE:",
                TEMP_MODE, "\tALPHA:", ALPHA, "\n")

            for _ in range(10):
                start = time.time()

                state, cost, states, costs = annealing(problem,random_start, cost_function, get_neighbour,
                                                        acceptance_probability, updateTemperature, NUM_ITERATIONS, DEBUG)
                
                duration = str(time.time() - start)[0:6]
                print('time:',duration , "\tcost:",cost)
                answers.append((state, cost, duration))

            printResult(answers)

            if EXCEl_WRITE:
                writeResultToExcel(filename, answers, myRow)
                myRow += 1
